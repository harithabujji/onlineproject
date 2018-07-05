import MySQLdb as mdb
import _mysql
import click
import sys

import os,django

os.environ['DJANGO_SETTINGS_MODULE'] = 'onlineproject.settings'
django.setup()

from onlineapp.models import *
from onlineproject import settings
from openpyxl import load_workbook

newDb = settings.DATABASES['default'].get("NAME")

@click.group()
def My_SQL_queries():
    """ supports some MySQLdb queries"""
    pass

@My_SQL_queries.command()

def createdb():
    newDb = settings.DATABASES['default'].get("NAME")
    print(newDb)
    con = mdb.connect("localhost", "root", "haritha123", newDb)
    cur = con.cursor()
    if con is None:
        click.echo('Error occured while connecting to datase')
        sys.exit(1)
    click.echo('creating  a database..')
    cursor = con.cursor()
    cursor.execute(f"CREATE DATABASE IF NOT EXISTS {newDb}")
    con.commit()
    os.system('python manage.py makemigrations')
    os.system('python manage.py migrate')
    pass



@My_SQL_queries.command("populatecollegetable",short_help="imports data into table")
def populatecollegetable():
    try:
        db = mdb.connect("localhost", "root", "haritha123", newDb)
    except:
        print("Can't connect to database")
        return 0
        # If Connection Is Successful
    print("Connected")

    cursor = db.cursor()
    wb1 = load_workbook("students.xlsx")
    sheets = wb1.sheetnames
    s = sheets[2]
    sheet = wb1[s]
    r = sheet.max_row
    c = sheet.max_column
    l = []
    for i in range(1, r + 1):
        l.append([])
    for i in range(1, r + 1):
        for j in range(1, c + 1):
            e = sheet.cell(row=i, column=j)
            l[i - 1].append(e.value)

    for i in range(1, r):

        c = College(name=l[i][0], acronym=l[i][1], location=l[i][2], contact=l[i][3])
        c.save()

    try:
        db.commit()
        print("commited")
    except:
        # Rollback in case there is any error
        db.rollback()
        print("rollback")

    db.close()

@My_SQL_queries.command("clearcollegedb",short_help="clears from table")
def clearcollegedb():
    try:
        db = mdb.connect("localhost", "root", "haritha123", newDb)
    except:
        print("Can't connect to database")
        return 0
        # If Connection Is Successful
    print("Connected")

    cursor = db.cursor()
    wb1 = load_workbook("students.xlsx")
    sheets = wb1.sheetnames
    s = sheets[2]
    sheet = wb1[s]
    r = sheet.max_row
    c = sheet.max_column

    l = []
    for i in range(1, r + 1):
        l.append([])

    for i in range(1, r + 1):
        for j in range(1, c + 1):
            e = sheet.cell(row=i, column=j)
            l[i - 1].append(e.value)

    sql = "TRUNCATE table onlineapp_college"
    cursor.execute(sql)
    db.commit()
    db.close()
@My_SQL_queries.command("populatestudents",short_help="imports data into table")
def populatestudents():
    try:
        db = mdb.connect("localhost", "root", "haritha123", newDb)
    except:
        print("Can't connect to database")
        return 0
        # If Connection Is Successful
    print("Connected")

    cursor = db.cursor()
    wb1 = load_workbook("students.xlsx")
    sheets = wb1.sheetnames
    s = sheets[0]
    sheet = wb1[s]
    r = sheet.max_row
    c = sheet.max_column
    l = []
    for i in range(1, r + 1):
        l.append([])
    for i in range(1, r + 1):
        for j in range(1, c + 1):
            e = sheet.cell(row=i, column=j)
            l[i - 1].append(e.value)

    for i in range(1, r):

        c = Student(name=l[i][0], email=l[i][2],db_folder=l[i][3])
        c.save()

    try:
        db.commit()
        print("commited")
    except:
        # Rollback in case there is any error
        db.rollback()
        print("rollback")

    db.close()


if __name__ == '__main__':

    My_SQL_queries()
