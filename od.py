
from openpyxl import load_workbook
import click
import os
import django
import MySQLdb


host='localhost'
username = 'root'
password = 'haritha123'


os.environ.setdefault("DJANGO_SETTINGS_MODULE","onlineproject.settings")
django.setup()


from onlineapp.models import *

def findDBName(str):
    str=str.replace("ol2016_",'')
    str=str.replace("_mock",'')
    return str[str.find('_')+1:]


@click.group()
def cli():
    pass


@cli.command()
def createdb():
    db = MySQLdb.connect(host, username, password)
    cursor = db.cursor()
    sql = 'create database temp_db'
    cursor.execute(sql)
    db.commit()
    db.close()


@cli.command()
def dropdb():
    db = MySQLdb.connect(host, username, password)
    cursor = db.cursor()
    sql = 'create database temp_db'
    cursor.execute(sql)
    db.commit()
    db.close()

@cli.command()
def cleardata():
    pass


@cli.command()
@click.argument('f1',nargs=1)
@click.argument('f2',nargs=1)
def populatedb(f1,f2):
    acroDict = {}
    studentsWorkbook = load_workbook(f1)
    CollegeWorksheet = studentsWorkbook.get_sheet_by_name('Colleges')
    for i in range(1, CollegeWorksheet.max_row):
        collegeData = []
        for j in range(0, CollegeWorksheet.max_column):
            cell = CollegeWorksheet.cell(row=i + 1, column=j + 1)
            collegeData.append(cell.value)
        acroDict[collegeData[1]] = i
        c = College(name=collegeData[0], acronym=collegeData[1], location=collegeData[2], contact=collegeData[3])
        c.save()


    nameDict = {}
    studentWorksheet = studentsWorkbook.get_sheet_by_name('Current')
    for i in range(1, studentWorksheet.max_row):
        studentData = []
        for j in range(0, studentWorksheet.max_column):
            cell = studentWorksheet.cell(row=i + 1, column=j + 1)
            studentData.append(cell.value)
        nameDict[studentData[3].lower()] = i
        s = Student(name=studentData[0], college=College(acroDict[studentData[1]]), db_folder=studentData[3], email=studentData[2])
        s.save()



    markssWorkbook = load_workbook(f2)
    marksWorksheet = markssWorkbook.get_sheet_by_name('dumpedsheet')
    for i in range(1, marksWorksheet.max_row):
        marksData = []
        for j in range(0, marksWorksheet.max_column):
            cell = marksWorksheet.cell(row=i + 1, column=j + 1)
            if j == 0:
                marksData.append(cell.value)
            else:
                marksData.append(int(cell.value))
        marksData[0] = findDBName(marksData[0])
        m = MockTest1(student=Student(nameDict[marksData[0]]), problem1=marksData[1], problem2=marksData[2],
                      problem3=marksData[3], problem4=marksData[4], total=marksData[5])
        m.save()

if __name__=='__main__':
    cli()
